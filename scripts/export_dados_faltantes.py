"""
scripts/export_dados_faltantes.py
Exporta para Excel os dados incompletos das tabelas Bling para revisão e merge.

Uso:
    python scripts/export_dados_faltantes.py
    python scripts/export_dados_faltantes.py --output relatorios/faltantes_jun.xlsx

Sheets geradas:
    1. Contatos Ativos     — contatos com atividade financeira e campos incompletos
    2. IDs Sem Cadastro    — contact_ids em pedidos/NFe/contas sem entry em bling_contacts
    3. NF-e Sem Total      — NF-es que precisam de busca individual para preencher total/série
    4. Contas Pagar Incompletas — contas sem descrição e nome do fornecedor
"""

import argparse
import os
import sys
from datetime import date, datetime
from pathlib import Path

import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Side,
    Border,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

# ---------------------------------------------------------------------------
# Paleta
# ---------------------------------------------------------------------------
COLOR_HEADER_BG  = "1F4E79"   # azul escuro
COLOR_HEADER_FG  = "FFFFFF"
COLOR_MISSING    = "FFF2CC"   # amarelo — campo faltante
COLOR_ORPHAN     = "FCE4D6"   # laranja claro — ID órfão
COLOR_ROW_ALT    = "EBF3FB"   # azul claríssimo — zebra
COLOR_TOTAL_BG   = "D6E4F0"   # azul médio — linha de totais

FONT_NAME = "Arial"

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _db():
    return psycopg2.connect(
        host=os.environ["DB_HOST"],
        port=int(os.environ.get("DB_PORT", 5432)),
        dbname=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASSWORD"],
        cursor_factory=psycopg2.extras.RealDictCursor,
    )


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def _header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_FG, size=10)
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _data_cell(ws, row, col, value, *, missing=False, alt_row=False, fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    if missing:
        c.fill = PatternFill("solid", fgColor=COLOR_MISSING)
    elif alt_row:
        c.fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
    if fmt:
        c.number_format = fmt
    return c


def _freeze_and_filter(ws, row=2):
    ws.freeze_panes = ws.cell(row=row, column=1)
    ws.auto_filter.ref = ws.dimensions


def _add_legend(ws, start_row):
    ws.cell(row=start_row, column=1, value="Legenda:").font = Font(name=FONT_NAME, bold=True, size=9)
    y = PatternFill("solid", fgColor=COLOR_MISSING)
    o = PatternFill("solid", fgColor=COLOR_ORPHAN)
    for col, (fill, label) in enumerate([
        (y, "Campo faltante — preencher para merge"),
        (o, "ID sem cadastro em bling_contacts"),
    ], start=2):
        c = ws.cell(row=start_row, column=col, value=label)
        c.fill = fill
        c.font = Font(name=FONT_NAME, size=9)
        c.border = BORDER


# ---------------------------------------------------------------------------
# Sheet 1: Contatos Ativos com campos incompletos
# ---------------------------------------------------------------------------

SQL_CONTATOS = """
WITH pedidos AS (
    SELECT client_id::bigint AS cid, COUNT(*) AS n, SUM(total) AS v
    FROM bling_orders WHERE client_id ~ '^[0-9]+$'
    GROUP BY client_id
),
nfe AS (
    SELECT contact_id AS cid, COUNT(*) AS n, SUM(total) AS v
    FROM bling_nfe GROUP BY contact_id
),
receber AS (
    SELECT contact_id AS cid,
           COUNT(*) AS n,
           SUM(value) AS v,
           SUM(value) FILTER (WHERE status = '1') AS aberto,
           SUM(value) FILTER (WHERE status = '2') AS recebido
    FROM bling_contas_receber GROUP BY contact_id
),
pagar AS (
    SELECT supplier_id AS cid,
           COUNT(*) AS n,
           SUM(value) AS v,
           SUM(value) FILTER (WHERE status = '1') AS aberto,
           SUM(value) FILTER (WHERE status = '2') AS pago
    FROM bling_contas_pagar GROUP BY supplier_id
)
SELECT
    c.id,
    c.name,
    c.document,
    c.person_type,
    c.email,
    c.phone,
    c.city,
    c.state,
    COALESCE(p.n, 0)  AS qtd_pedidos,
    COALESCE(p.v, 0)  AS total_pedidos,
    COALESCE(n.n, 0)  AS qtd_nfe,
    COALESCE(n.v, 0)  AS total_nfe,
    COALESCE(r.n, 0)  AS qtd_receber,
    COALESCE(r.v, 0)  AS total_receber,
    COALESCE(r.aberto, 0)    AS receber_aberto,
    COALESCE(r.recebido, 0)  AS receber_recebido,
    COALESCE(pg.n, 0) AS qtd_pagar,
    COALESCE(pg.v, 0) AS total_pagar,
    COALESCE(pg.aberto, 0)   AS pagar_aberto,
    COALESCE(pg.pago, 0)     AS pagar_pago,
    COALESCE(r.v, 0) - COALESCE(pg.v, 0) AS saldo_liquido
FROM bling_contacts c
LEFT JOIN pedidos  p  ON p.cid = c.id
LEFT JOIN nfe      n  ON n.cid = c.id
LEFT JOIN receber  r  ON r.cid = c.id
LEFT JOIN pagar    pg ON pg.cid = c.id
WHERE (COALESCE(p.n,0) + COALESCE(n.n,0) + COALESCE(r.n,0) + COALESCE(pg.n,0)) > 0
ORDER BY (COALESCE(p.v,0) + COALESCE(r.v,0)) DESC
"""

HEADERS_CONTATOS = [
    ("ID Bling",         10, None),
    ("Nome",             30, None),
    ("Documento",        18, None),
    ("Tipo",              7, None),
    ("E-mail",           26, None),
    ("Telefone",         14, None),
    ("Cidade",           16, None),
    ("UF",                5, None),
    ("Qtd Pedidos",      11, "0"),
    ("Total Pedidos (R$)", 18, '#,##0.00'),
    ("Qtd NF-e",         10, "0"),
    ("Total NF-e (R$)",  16, '#,##0.00'),
    ("Qtd Receber",      11, "0"),
    ("Total Receber (R$)", 18, '#,##0.00'),
    ("Rec. Aberto (R$)", 16, '#,##0.00'),
    ("Rec. Recebido (R$)", 18, '#,##0.00'),
    ("Qtd Pagar",        10, "0"),
    ("Total Pagar (R$)", 16, '#,##0.00'),
    ("Pagar Aberto (R$)", 16, '#,##0.00'),
    ("Pagar Pago (R$)",  16, '#,##0.00'),
    ("Saldo Líquido (R$)", 18, '#,##0.00;(#,##0.00)'),
]

# colunas de texto que podem estar null/vazio (índice 0-based nos dados)
MISSING_COLS_CONTATOS = {
    "document": 2,
    "person_type": 3,
    "email": 4,
    "phone": 5,
    "city": 6,
    "state": 7,
}


def build_sheet_contatos(ws, rows):
    ws.title = "Contatos Ativos"
    ws.row_dimensions[1].height = 30

    for col, (label, width, _) in enumerate(HEADERS_CONTATOS, start=1):
        _header_cell(ws, 1, col, label, width)

    fields = ["id","name","document","person_type","email","phone","city","state",
              "qtd_pedidos","total_pedidos","qtd_nfe","total_nfe",
              "qtd_receber","total_receber","receber_aberto","receber_recebido",
              "qtd_pagar","total_pagar","pagar_aberto","pagar_pago","saldo_liquido"]

    for r_idx, row in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, (field, (_, _, fmt)) in enumerate(zip(fields, HEADERS_CONTATOS), start=1):
            val = row[field]
            # null ou string vazia em coluna textual → faltante
            is_text_col = field in MISSING_COLS_CONTATOS
            missing = is_text_col and (val is None or str(val).strip() == "")
            align = "right" if fmt and "0" in fmt else "left"
            _data_cell(ws, r_idx, c_idx, val, missing=missing, alt_row=alt, fmt=fmt, align=align)

    _freeze_and_filter(ws)

    # Totais
    total_row = len(rows) + 2
    ws.cell(total_row, 1, "TOTAL").font = Font(name=FONT_NAME, bold=True, size=9)
    for c_idx, (_, _, fmt) in enumerate(HEADERS_CONTATOS, start=1):
        if fmt and "0" in fmt and c_idx > 8:
            letter = get_column_letter(c_idx)
            c = ws.cell(total_row, c_idx, f"=SUM({letter}2:{letter}{total_row-1})")
            c.font = Font(name=FONT_NAME, bold=True, size=9)
            c.fill = PatternFill("solid", fgColor=COLOR_TOTAL_BG)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            c.border = BORDER

    _add_legend(ws, total_row + 2)
    return len(rows)


# ---------------------------------------------------------------------------
# Sheet 2: IDs órfãos (contact_id sem entry em bling_contacts)
# ---------------------------------------------------------------------------

SQL_ORFAOS = """
SELECT 'nfe'             AS origem,
       n.contact_id      AS contact_id,
       n.contact_name    AS nome_na_origem,
       COUNT(*)          AS ocorrencias,
       MIN(n.issue_date) AS primeira_data,
       MAX(n.issue_date) AS ultima_data,
       SUM(n.total)      AS valor_total
FROM bling_nfe n
LEFT JOIN bling_contacts c ON c.id = n.contact_id
WHERE c.id IS NULL AND n.contact_id IS NOT NULL
GROUP BY n.contact_id, n.contact_name

UNION ALL

SELECT 'contas_pagar'    AS origem,
       cp.supplier_id    AS contact_id,
       cp.supplier_name  AS nome_na_origem,
       COUNT(*)          AS ocorrencias,
       MIN(cp.due_date)  AS primeira_data,
       MAX(cp.due_date)  AS ultima_data,
       SUM(cp.value)     AS valor_total
FROM bling_contas_pagar cp
LEFT JOIN bling_contacts c ON c.id = cp.supplier_id
WHERE c.id IS NULL AND cp.supplier_id IS NOT NULL
GROUP BY cp.supplier_id, cp.supplier_name

UNION ALL

SELECT 'orders'          AS origem,
       o.client_id::bigint AS contact_id,
       o.client_name     AS nome_na_origem,
       COUNT(*)          AS ocorrencias,
       MIN(o.created_at::date) AS primeira_data,
       MAX(o.created_at::date) AS ultima_data,
       SUM(o.total)      AS valor_total
FROM bling_orders o
LEFT JOIN bling_contacts c ON c.id = o.client_id::bigint
WHERE c.id IS NULL AND o.client_id ~ '^[0-9]+$'
GROUP BY o.client_id, o.client_name

ORDER BY origem, contact_id
"""

HEADERS_ORFAOS = [
    ("Origem",           16),
    ("ID Bling",         14),
    ("Nome (na origem)", 32),
    ("Ocorrências",      13),
    ("Primeira Data",    14),
    ("Última Data",      14),
    ("Valor Total (R$)", 18),
]


def build_sheet_orfaos(ws, rows):
    ws.title = "IDs Sem Cadastro"
    ws.row_dimensions[1].height = 28

    for col, (label, width) in enumerate(HEADERS_ORFAOS, start=1):
        _header_cell(ws, 1, col, label, width)

    fields = ["origem","contact_id","nome_na_origem","ocorrencias","primeira_data","ultima_data","valor_total"]
    for r_idx, row in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, field in enumerate(fields, start=1):
            val = row[field]
            fmt = '#,##0.00' if field == "valor_total" else ("0" if field == "ocorrencias" else None)
            align = "right" if fmt else "left"
            c = _data_cell(ws, r_idx, c_idx, val, alt_row=alt, fmt=fmt, align=align)
            c.fill = PatternFill("solid", fgColor=COLOR_ORPHAN)

    _freeze_and_filter(ws)
    return len(rows)


# ---------------------------------------------------------------------------
# Sheet 3: NF-e sem total/série
# ---------------------------------------------------------------------------

SQL_NFE = """
SELECT
    id,
    numero,
    serie,
    situation,
    contact_id,
    contact_name,
    total,
    issue_date,
    updated_at
FROM bling_nfe
WHERE total IS NULL OR serie IS NULL
ORDER BY issue_date DESC NULLS LAST
"""

HEADERS_NFE = [
    ("ID Bling",      14),
    ("Número NF",     12),
    ("Série",          8),
    ("Situação",      10),
    ("Contact ID",    14),
    ("Nome Contato",  28),
    ("Total (R$)",    14),
    ("Data Emissão",  14),
    ("Atualizado em", 18),
]


def build_sheet_nfe(ws, rows):
    ws.title = "NF-e Sem Total"
    ws.row_dimensions[1].height = 28

    for col, (label, width) in enumerate(HEADERS_NFE, start=1):
        _header_cell(ws, 1, col, label, width)

    fields = ["id","numero","serie","situation","contact_id","contact_name","total","issue_date","updated_at"]
    fmt_map = {"total": '#,##0.00'}

    for r_idx, row in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, field in enumerate(fields, start=1):
            val = row[field]
            missing = val is None and field in ("serie", "total")
            fmt = fmt_map.get(field)
            align = "right" if fmt else "left"
            _data_cell(ws, r_idx, c_idx, val, missing=missing, alt_row=alt, fmt=fmt, align=align)

    _freeze_and_filter(ws)
    return len(rows)


# ---------------------------------------------------------------------------
# Sheet 4: Contas Pagar incompletas
# ---------------------------------------------------------------------------

SQL_PAGAR = """
SELECT
    id,
    description,
    supplier_id,
    supplier_name,
    due_date,
    value,
    status,
    competency,
    category,
    updated_at
FROM bling_contas_pagar
WHERE description IS NULL OR supplier_name IS NULL OR category IS NULL
ORDER BY due_date DESC NULLS LAST
"""

HEADERS_PAGAR = [
    ("ID Bling",       14),
    ("Descrição",      32),
    ("Supplier ID",    14),
    ("Fornecedor",     28),
    ("Vencimento",     14),
    ("Valor (R$)",     14),
    ("Status",         10),
    ("Competência",    14),
    ("Categoria",      20),
    ("Atualizado em",  18),
]


def build_sheet_pagar(ws, rows):
    ws.title = "Contas Pagar Incompletas"
    ws.row_dimensions[1].height = 28

    for col, (label, width) in enumerate(HEADERS_PAGAR, start=1):
        _header_cell(ws, 1, col, label, width)

    fields = ["id","description","supplier_id","supplier_name","due_date",
              "value","status","competency","category","updated_at"]
    missing_fields = {"description", "supplier_name", "category"}
    fmt_map = {"value": '#,##0.00'}
    STATUS = {"1": "Aberto", "2": "Pago", "5": "Parcial"}

    for r_idx, row in enumerate(rows, start=2):
        alt = (r_idx % 2 == 0)
        for c_idx, field in enumerate(fields, start=1):
            val = row[field]
            if field == "status" and val:
                val = STATUS.get(str(val), val)
            missing = field in missing_fields and (val is None or str(val).strip() == "")
            fmt = fmt_map.get(field)
            align = "right" if fmt else "left"
            _data_cell(ws, r_idx, c_idx, val, missing=missing, alt_row=alt, fmt=fmt, align=align)

    _freeze_and_filter(ws)
    return len(rows)


# ---------------------------------------------------------------------------
# Capa / Instruções
# ---------------------------------------------------------------------------

def build_sheet_capa(ws, counts: dict, gerado_em: datetime):
    ws.title = "Instruções"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    title = ws.cell(1, 1, "Relatório: Dados Faltantes Bling")
    title.font = Font(name=FONT_NAME, bold=True, size=14, color=COLOR_HEADER_BG)
    ws.merge_cells("A1:B1")
    title.alignment = Alignment(horizontal="center")

    info = [
        ("Gerado em",          gerado_em.strftime("%d/%m/%Y %H:%M")),
        ("",""),
        ("Sheet",              "Conteúdo"),
        ("Contatos Ativos",    "Contatos com atividade financeira. Campos amarelos = faltante."),
        ("IDs Sem Cadastro",   "contact_id encontrado em pedidos/NFe/contas mas ausente em bling_contacts."),
        ("NF-e Sem Total",     "NF-es sem valor total e série — buscar via GET /nfe/{id}."),
        ("Contas Pagar Inc.",  "Contas sem descrição/fornecedor — buscar via GET /contas/pagar/{id}."),
        ("",""),
        ("Contatos Ativos",    f"{counts.get('contatos', 0):,} linhas"),
        ("IDs Sem Cadastro",   f"{counts.get('orfaos', 0):,} linhas"),
        ("NF-e Sem Total",     f"{counts.get('nfe', 0):,} linhas"),
        ("Contas Pagar Inc.",  f"{counts.get('pagar', 0):,} linhas"),
        ("",""),
        ("Merge — como usar",  ""),
        ("1.",                 "Abrir a sheet relevante."),
        ("2.",                 "Campos com fundo amarelo estão nulos no banco."),
        ("3.",                 "Preencher manualmente OU rodar enriquecimento via API."),
        ("4.",                 "Para reimportar: usar blng_fetcher/merge_from_xlsx.py (a criar)."),
        ("5.",                 "Rodar este script novamente para atualizar o relatório."),
    ]

    hdr = Font(name=FONT_NAME, bold=True, size=10, color=COLOR_HEADER_FG)
    norm = Font(name=FONT_NAME, size=10)

    for r_idx, (k, v) in enumerate(info, start=3):
        ck = ws.cell(r_idx, 1, k)
        cv = ws.cell(r_idx, 2, v)
        is_section = r_idx in (5, 11, 15)
        if is_section:
            ck.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            cv.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
            ck.font = hdr
            cv.font = hdr
        else:
            ck.font = norm
            cv.font = norm


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(output_path: str):
    print("Conectando ao banco...")
    conn = _db()

    try:
        with conn.cursor() as cur:
            cur.execute(SQL_CONTATOS)
            rows_contatos = cur.fetchall()
            cur.execute(SQL_ORFAOS)
            rows_orfaos = cur.fetchall()
            cur.execute(SQL_NFE)
            rows_nfe = cur.fetchall()
            cur.execute(SQL_PAGAR)
            rows_pagar = cur.fetchall()
    finally:
        conn.close()

    print(f"Contatos ativos: {len(rows_contatos)}")
    print(f"IDs órfãos:      {len(rows_orfaos)}")
    print(f"NF-e sem total:  {len(rows_nfe)}")
    print(f"Contas pagar:    {len(rows_pagar)}")

    wb = Workbook()
    wb.remove(wb.active)

    now = datetime.now()
    ws_capa = wb.create_sheet("Instruções")
    ws_cont = wb.create_sheet("Contatos Ativos")
    ws_orf  = wb.create_sheet("IDs Sem Cadastro")
    ws_nfe  = wb.create_sheet("NF-e Sem Total")
    ws_pag  = wb.create_sheet("Contas Pagar Incompletas")

    n_cont = build_sheet_contatos(ws_cont, rows_contatos)
    n_orf  = build_sheet_orfaos(ws_orf, rows_orfaos)
    n_nfe  = build_sheet_nfe(ws_nfe, rows_nfe)
    n_pag  = build_sheet_pagar(ws_pag, rows_pagar)
    build_sheet_capa(ws_capa, {
        "contatos": n_cont,
        "orfaos":   n_orf,
        "nfe":      n_nfe,
        "pagar":    n_pag,
    }, now)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\nArquivo gerado: {out.resolve()}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Exporta dados faltantes Bling para Excel")
    default_name = f"dados_faltantes_{date.today().strftime('%Y%m%d')}.xlsx"
    parser.add_argument(
        "--output", default=str(ROOT / "relatorios" / default_name),
        help=f"Caminho do arquivo de saída (default: relatorios/{default_name})",
    )
    args = parser.parse_args()
    main(args.output)
