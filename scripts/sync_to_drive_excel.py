"""
scripts/sync_to_drive_excel.py
Gera .xlsx e .xml com as mesmas entidades do sync_to_sheets.py (uma aba/
elemento por entidade) e sobrescreve dois arquivos fixos numa pasta do
Google Drive.

Canal adicional ao Google Sheets (nao substitui) -- ver sync_to_sheets.py
para a planilha ao vivo. Aqui os arquivos sao estaticos: cada execucao
gera os dois do zero localmente (a partir do MESMO fetch, sem consultar o
banco em dobro) e substitui o conteudo na pasta.

.env necessario (reaproveita o mesmo do sync_to_sheets.py):
    DB_HOST, DB_PORT, DB_NAME, DB_USER, DB_PASSWORD
    GSHEETS_CREDENTIALS=path/to/service_account.json (mesmo service account)

Uso:
    python scripts/sync_to_drive_excel.py
    python scripts/sync_to_drive_excel.py --folder-id 1MzCuJ0X_...
    python scripts/sync_to_drive_excel.py --formats xlsx     # so' um dos dois
    python scripts/sync_to_drive_excel.py --dry-run          # gera local, nao envia
"""
import argparse
import logging
import os
import re
import sys
import time
import unicodedata
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DEFAULT_FOLDER_ID = "0ACb496WJWjjZUk9PVA"  # Shared Drive "BLING - Drivers compartilhados"
DEFAULT_XLSX_FILENAME = "Bling - Relatorio.xlsx"
DEFAULT_XML_FILENAME = "Bling - Relatorio.xml"
XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XML_MIMETYPE = "application/xml"

# Nomes de aba do Excel tem limite de 31 chars e nao aceitam alguns caracteres
_INVALID_SHEET_CHARS = set(r"[]:*?/\\")

# XML 1.0 (base do .xlsx tambem) rejeita caracteres de controle fora de
# tab/CR/LF; o Sheets tolera esses valores (varios registros do Bling trazem
# lixo desse tipo em campos de texto), entao os dois formatos precisam
# sanitizar antes de gravar.
_ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]")


def _sanitize_text(value: str) -> str:
    return _ILLEGAL_XML_CHARS_RE.sub("", value)


def _sanitize_cell(value):
    if isinstance(value, str):
        return _sanitize_text(value)
    return value


def _cell_to_text(value) -> str | None:
    """Mesmo formato usado no Sheets (_fmt em sync_to_sheets.py), como texto."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        f = float(value)
        return str(int(f)) if f == int(f) else str(f)
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return _sanitize_text(str(value))


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = "".join(c for c in name if c not in _INVALID_SHEET_CHARS)[:31]
    base, i = cleaned, 1
    while cleaned in used:
        suffix = f"~{i}"
        cleaned = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(cleaned)
    return cleaned


def _xml_safe_tag(header: str, used: set[str]) -> str:
    """Converte um header em portugues ('Nº Pedido', 'Valor (R$)') numa tag
    XML valida: sem acento, sem espaco/parenteses, nao comeca com digito."""
    ascii_name = unicodedata.normalize("NFKD", header).encode("ascii", "ignore").decode()
    tag = re.sub(r"\W+", "_", ascii_name).strip("_") or "campo"
    if tag[0].isdigit():
        tag = f"_{tag}"
    base, i = tag, 1
    while tag in used:
        tag = f"{base}_{i}"
        i += 1
    used.add(tag)
    return tag


def _load_all_entities(entity_keys: list[str]) -> dict:
    """Busca cada entidade UMA vez; xlsx e xml sao gerados do mesmo resultado."""
    from scripts.sync_to_sheets import ENTITY_MAP, _fetch

    data = {}
    for key in entity_keys:
        entity = ENTITY_MAP[key]
        logger.info("[%s] Buscando dados...", entity.tab)
        rows = _fetch(entity.sql)
        logger.info("[%s] %s linhas.", entity.tab, len(rows))
        data[key] = (entity, rows)
    return data


def build_workbook(loaded: dict, out_path: Path) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # remove a aba default vazia
    used_names: set[str] = set()
    total_rows = 0

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for entity, rows in loaded.values():
        total_rows += len(rows)
        ws = wb.create_sheet(_safe_sheet_name(entity.tab, used_names))
        ws.append(entity.headers)
        for row in rows:
            ws.append([_sanitize_cell(v) for v in row])

        for col_idx in range(1, len(entity.headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
        ws.freeze_panes = "A2"
        for col_idx, header in enumerate(entity.headers, start=1):
            width = min(max(len(header) + 2, 10), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)
    logger.info("Workbook salvo em %s (%s abas, %s linhas no total).",
               out_path, len(loaded), total_rows)
    return total_rows


def build_xml(loaded: dict, out_path: Path) -> int:
    root = ET.Element("blingExport", attrib={
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
    })
    total_rows = 0

    for key, (entity, rows) in loaded.items():
        entity_el = ET.SubElement(root, "entity", attrib={"key": key, "tab": entity.tab})
        used_tags: set[str] = set()
        tags = [_xml_safe_tag(h, used_tags) for h in entity.headers]
        total_rows += len(rows)

        for row in rows:
            row_el = ET.SubElement(entity_el, "row")
            for tag, value in zip(tags, row):
                text = _cell_to_text(value)
                field_el = ET.SubElement(row_el, tag)
                if text is not None:
                    field_el.text = text

    tree = ET.ElementTree(root)
    ET.indent(tree, space="  ")
    tree.write(out_path, encoding="utf-8", xml_declaration=True)
    logger.info("XML salvo em %s (%s entidades, %s linhas no total).",
               out_path, len(loaded), total_rows)
    return total_rows


def _build_drive_service(credentials_path: str | None):
    import httplib2
    from google.oauth2.service_account import Credentials
    from google_auth_httplib2 import AuthorizedHttp
    from googleapiclient.discovery import build

    creds_file = credentials_path or os.environ.get("GSHEETS_CREDENTIALS")
    creds = Credentials.from_service_account_file(
        creds_file, scopes=["https://www.googleapis.com/auth/drive"])
    # timeout maior + cache_discovery=False: a rede local tem instabilidade
    # transitoria (visto tambem no Postgres/Bling nesta mesma maquina); o
    # timeout default do httplib2 e' curto demais e derruba a chamada antes
    # de a conexao (que geralmente completa, so' que devagar) se estabelecer.
    http = AuthorizedHttp(creds, http=httplib2.Http(timeout=30))
    return build("drive", "v3", http=http, cache_discovery=False)


def upload_or_replace(service, folder_id: str, filename: str, local_path: Path,
                      mimetype: str) -> str:
    from googleapiclient.http import MediaFileUpload

    # supportsAllDrives: necessario para o Drive API enxergar/gravar em Shared
    # Drives (sem isso, chamadas em itens de um Shared Drive falham/ignoram
    # silenciosamente). Inofensivo para pastas normais do Meu Drive tambem.
    query = (
        f"name = '{filename}' and '{folder_id}' in parents "
        "and trashed = false"
    )
    existing = service.files().list(
        q=query, fields="files(id,name)", spaces="drive",
        includeItemsFromAllDrives=True, supportsAllDrives=True,
        corpora="allDrives",
    ).execute().get("files", [])

    media = MediaFileUpload(str(local_path), mimetype=mimetype, resumable=True)

    if existing:
        file_id = existing[0]["id"]
        for attempt in range(3):
            try:
                service.files().update(
                    fileId=file_id, media_body=media, supportsAllDrives=True,
                ).execute()
                logger.info("Arquivo existente atualizado (id=%s, %s).", file_id, filename)
                return file_id
            except Exception as exc:  # noqa: BLE001
                if attempt == 2:
                    raise
                logger.warning("Falha ao atualizar (%s); tentando de novo em 5s...", exc)
                time.sleep(5)
    else:
        # Nota: se folder_id for uma pasta comum do Meu Drive (nao um Shared
        # Drive), esta chamada de CRIACAO ainda vai falhar com
        # storageQuotaExceeded -- contas de servico so tem cota propria em
        # Shared Drives. Update (acima) funciona em qualquer lugar contanto
        # que o arquivo ja exista e a conta tenha permissao de edicao.
        metadata = {"name": filename, "parents": [folder_id]}
        created = service.files().create(
            body=metadata, media_body=media, fields="id", supportsAllDrives=True,
        ).execute()
        file_id = created["id"]
        logger.info("Arquivo novo criado (id=%s, %s).", file_id, filename)
        return file_id


def main(folder_id: str, xlsx_filename: str, xml_filename: str,
         credentials_path: str | None, entity_keys: list[str],
         formats: set[str], dry_run: bool):
    from scripts.sync_to_sheets import ENTITY_MAP

    unknown = set(entity_keys) - set(ENTITY_MAP)
    if unknown:
        raise SystemExit(f"Entidades desconhecidas: {unknown}")

    out_dir = ROOT / "relatorios"
    out_dir.mkdir(parents=True, exist_ok=True)

    loaded = _load_all_entities(entity_keys)

    outputs: list[tuple[Path, str]] = []
    if "xlsx" in formats:
        xlsx_path = out_dir / xlsx_filename
        build_workbook(loaded, xlsx_path)
        outputs.append((xlsx_path, XLSX_MIMETYPE))
    if "xml" in formats:
        xml_path = out_dir / xml_filename
        build_xml(loaded, xml_path)
        outputs.append((xml_path, XML_MIMETYPE))

    if dry_run:
        logger.info("--dry-run: arquivo(s) gerado(s) localmente em %s, nada enviado.", out_dir)
        return

    logger.info("Conectando ao Google Drive...")
    service = _build_drive_service(credentials_path)
    for path, mimetype in outputs:
        upload_or_replace(service, folder_id, path.name, path, mimetype)
    logger.info("Concluido.")


if __name__ == "__main__":
    from scripts.sync_to_sheets import ENTITY_MAP  # noqa: E402

    parser = argparse.ArgumentParser(
        description="Bling -> .xlsx/.xml sincronizados numa pasta do Google Drive")
    parser.add_argument("--folder-id", default=DEFAULT_FOLDER_ID)
    parser.add_argument("--xlsx-filename", default=DEFAULT_XLSX_FILENAME)
    parser.add_argument("--xml-filename", default=DEFAULT_XML_FILENAME)
    parser.add_argument(
        "--formats", default="xlsx,xml",
        help="Formatos a gerar/enviar, separados por virgula (default: xlsx,xml)",
    )
    parser.add_argument(
        "--credentials", default=os.environ.get("GSHEETS_CREDENTIALS"),
        help="Caminho do service_account.json (ou GSHEETS_CREDENTIALS no .env)",
    )
    parser.add_argument(
        "--entity", action="append", dest="entities",
        help="Restringe a entidades especificas (repita p/ mais de uma); default: todas",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    keys = args.entities or list(ENTITY_MAP.keys())
    formats = {f.strip() for f in args.formats.split(",") if f.strip()}
    main(
        folder_id=args.folder_id, xlsx_filename=args.xlsx_filename,
        xml_filename=args.xml_filename, credentials_path=args.credentials,
        entity_keys=keys, formats=formats, dry_run=args.dry_run,
    )
