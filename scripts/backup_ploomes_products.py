"""
Cria um backup completo de todos os produtos do Ploomes nos formatos JSON e CSV/Excel.

Uso:
    python scripts/backup_ploomes_products.py
"""

import os
import sys
import json
import csv
from datetime import datetime
from pathlib import Path

# Garante que a raiz do projeto está no path de importação
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.config import get_settings
from app.clients.ploomes import PloomesClient

# Tentativa de importar openpyxl para gerar planilha Excel bonita. Se não existir, usa CSV como fallback.
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def get_other_property_value(other_properties, field_key):
    if not other_properties or not field_key:
        return ""
    for prop in other_properties:
        if prop.get("FieldKey") == field_key:
            return (
                prop.get("StringValue") or
                prop.get("IntegerValue") or
                prop.get("BigStringValue") or
                ""
            )
    return ""

def main():
    print("=== INICIANDO BACKUP DE PRODUTOS DO PLOOMES ===")
    
    settings = get_settings()
    if not settings.ploomes_user_key:
        print("Erro: PLOOMES_USER_KEY não está configurado no arquivo .env!")
        sys.exit(1)
        
    client = PloomesClient(settings)
    
    # Criar pasta de backups se não existir
    backup_dir = ROOT / "backups"
    backup_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = backup_dir / f"ploomes_products_backup_{timestamp}.json"
    xlsx_path = backup_dir / f"ploomes_products_backup_{timestamp}.xlsx"
    csv_path = backup_dir / f"ploomes_products_backup_{timestamp}.csv"
    
    products = []
    print("Buscando produtos do Ploomes paginado (100 por vez)...")
    
    count = 0
    try:
        for product in client.iter_products(page_size=100):
            other_props = product.get("OtherProperties", [])
            
            # Extrai os campos customizados conforme mapeado nas configurações do app
            fabricante = get_other_property_value(other_props, settings.ploomes_field_fabricante)
            partnumber = get_other_property_value(other_props, settings.ploomes_field_partnumber)
            ncm = get_other_property_value(other_props, settings.ploomes_field_ncm)
            descricao = get_other_property_value(other_props, settings.ploomes_field_descricao)
            
            flat_product = {
                "Id": product.get("Id"),
                "Code": product.get("Code"),
                "Name": product.get("Name"),
                "UnitPrice": product.get("UnitPrice"),
                "Suspended": product.get("Suspended"),
                "Fabricante": fabricante,
                "Partnumber": partnumber,
                "NCM": ncm,
                "Descricao": descricao,
                "Raw": product # mantém dados brutos
            }
            products.append(flat_product)
            count += 1
            if count % 100 == 0:
                print(f"  -> {count} produtos baixados...")
    except Exception as exc:
        print(f"\nErro ao buscar produtos do Ploomes: {exc}")
        sys.exit(1)
        
    print(f"\nBusca concluída! Total de produtos encontrados: {count}")
    
    if count == 0:
        print("Nenhum produto cadastrado no Ploomes. Não há necessidade de backup.")
        return
        
    # 1. Salvar backup em formato JSON (contém os metadados completos brutos)
    print(f"Salvando backup JSON em: backups/{json_path.name}")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump([p["Raw"] for p in products], f, indent=2, ensure_ascii=False)
        
    # 2. Salvar backup de visualização (Excel ou CSV)
    if OPENPYXL_AVAILABLE:
        print(f"Salvando backup Excel (.xlsx) em: backups/{xlsx_path.name}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos Ploomes"
        
        # Cabeçalhos do Excel
        headers = ["ID Ploomes", "Código (SKU)", "Nome", "Preço Unitário", "Inativo (Suspended)", "Fabricante", "Partnumber", "NCM", "Descrição Completa"]
        
        # Estilização
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        for row_idx, p in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=p["Id"]).alignment = center_align
            ws.cell(row=row_idx, column=2, value=p["Code"]).alignment = center_align
            ws.cell(row=row_idx, column=3, value=p["Name"]).alignment = left_align
            
            price_cell = ws.cell(row=row_idx, column=4, value=p["UnitPrice"])
            price_cell.number_format = '"R$"#,##0.00'
            price_cell.alignment = Alignment(horizontal="right", vertical="center")
            
            suspended_text = "Sim" if p["Suspended"] else "Não"
            ws.cell(row=row_idx, column=5, value=suspended_text).alignment = center_align
            
            ws.cell(row=row_idx, column=6, value=p["Fabricante"]).alignment = left_align
            ws.cell(row=row_idx, column=7, value=p["Partnumber"]).alignment = center_align
            ws.cell(row=row_idx, column=8, value=p["NCM"]).alignment = center_align
            ws.cell(row=row_idx, column=9, value=p["Descricao"]).alignment = left_align
            
            for col_idx in range(1, 10):
                ws.cell(row=row_idx, column=col_idx).border = thin_border
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        wb.save(xlsx_path)
    else:
        print("\nAviso: Biblioteca 'openpyxl' não encontrada.")
        print(f"Salvando backup estruturado em CSV como fallback em: backups/{csv_path.name}")
        
        # Cabeçalhos do CSV
        headers = ["ID Ploomes", "Codigo (SKU)", "Nome", "Preco Unitario", "Inativo (Suspended)", "Fabricante", "Partnumber", "NCM", "Descricao Completa"]
        
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            for p in products:
                writer.writerow([
                    p["Id"],
                    p["Code"],
                    p["Name"],
                    f"{p['UnitPrice']:.2f}".replace(".", ","),
                    "Sim" if p["Suspended"] else "Não",
                    p["Fabricante"],
                    p["Partnumber"],
                    p["NCM"],
                    p["Descricao"]
                ])
                
        print("\nDica: Se você deseja gerar uma planilha em Excel formatada (.xlsx), instale o pacote openpyxl:")
        print("  .\\.venv\\Scripts\\pip install openpyxl")
        
    print("\n=== BACKUP CONCLUÍDO COM SUCESSO! ===")

if __name__ == "__main__":
    main()
